import openpyxl
import os
from functools import reduce
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


class ExcelHandler:
    titles = {
        'sample': ['農戶編號', '調查姓名', '電話', '地址', '出生年', '原層別', '連結編號'],
        'household': ['[戶籍檔]', '出生年', '關係'],
        'crop_sbdy': ['[轉作補貼]', '項目', '作物名稱', '期別'],
        'disaster': ['[災害]', '項目', '災害', '核定作物', '核定面積'],
        'crops_name': '[106y-107y作物]',
        'fir_half': ['', '一月', '二月', '三月', '四月', '五月', '六月'],
        'sec_half': ['七月', '八月', '九月', '十月', '十一月', '十二月'],
        'hire': '[106每月常僱員工]',
        'lack_situation': '[106勞動力短缺情形]',
        'lack': '[106短缺常僱員工]',
        'short_lack': '[106短缺臨時僱工]',
    }

    def __init__(self, county, path):
        self.__county = county
        self.__path = path
        self.__col_index = 1
        self.__row_index = 1
        self.__wb = openpyxl.Workbook()
        self.__sheet = self.__wb.active
        self.__crop_set = set()

        self.__set_column_width()

    @property
    def column_index(self):
        return self.__col_index

    @column_index.setter
    def column_index(self, i):
        if i == -1:
            self.__col_index = 1
        else:
            self.__col_index += i

    @property
    def row_index(self):
        return self.__row_index

    @row_index.setter
    def row_index(self, i):
        if i == -1:
            self.__row_index = 1
        else:
            self.__row_index += i

    def __set_column_width(self):
        width = list(map(lambda x: x * 1.054, [20.29, 9.29, 16.29, 29.29, 9.29, 11.29, 11.29, 11.29, 11.29]))
        for i in range(1, len(width) + 1):
            self.__sheet.column_dimensions[get_column_letter(i)].width = width[i - 1]
        self.row_index = 1

    def __set_title(self, title, start=1):
        for index, _title in enumerate(self.titles[title], start=start):
            self.__sheet.cell(column=index, row=self.row_index).value = _title
        self.row_index = 1

    def __set_sample_base_data(self, data):
        self.__set_title('sample')
        value_list = [data['farmer_num'], data['name'], data['tel'],
                      data['addr'], data['birthday'], data['layer'], data['link_num']]

        for index, val in enumerate(value_list, start=1):
            self.__sheet.cell(column=index, row=self.row_index).value = val
            self.__sheet.cell(column=index, row=self.row_index).alignment = Alignment(wrap_text=True)
        self.row_index = 1
        self.__sheet.cell(column=self.column_index, row=self.row_index).value = ' ' + '-' * 206 + ' '
        self.row_index = 1

    def __set_household_data(self, members):
        self.__set_title('household')
        if not members:
            return
        for person in members:
            for index, val in enumerate(person, start=2):
                self.__sheet.cell(column=index, row=self.row_index).value = val
                self.__sheet.cell(column=index, row=self.row_index).alignment = Alignment(horizontal='left')

    def __set_crop_sbdy_data(self, crop_sbdy_list):
        if not crop_sbdy_list:
            return
        self.row_index = 2
        self.__set_title('crop_sbdy')
        self.__crop_set = crop_name = {i[0] for i in crop_sbdy_list}

        for index, val in enumerate(crop_name, start=1):
            if index >= 2:
                self.row_index = 1
            self.__sheet.cell(column=2, row=self.row_index).value = index
            self.__sheet.cell(column=3, row=self.row_index).value = val
            self.__sheet.cell(column=4, row=self.row_index).value = '1'

            if len(val) > 8:
                self.__sheet.cell(column=3, row=self.row_index).alignment = Alignment(wrap_text=True)

    def __set_disaster_data(self, disaster_list):
        if not disaster_list:
            return
        _dict = {}
        self.row_index = 2
        for d in disaster_list:
            key = (d[0], d[1])
            if key not in _dict:
                _dict[key] = float(d[2])
            else:
                _dict[key] = _dict.get(key) + float(d[2])
        self.__set_title('disaster')

        for index, _tuple in enumerate(_dict.items(), start=1):
            if index >= 2:
                self.row_index = 1
            self.__crop_set.add(_tuple[0][1])
            self.__sheet.cell(column=2, row=self.row_index).value = index
            self.__sheet.cell(column=3, row=self.row_index).value = _tuple[0][0]
            self.__sheet.cell(column=4, row=self.row_index).value = _tuple[0][1]
            self.__sheet.cell(column=5, row=self.row_index).value = _tuple[1]

    def __set_crops_name(self):
        if not self.__crop_set:
            return
        self.row_index = 2
        crops = reduce(lambda a, b: a + ', ' + b, self.__crop_set)
        self.__sheet.cell(column=1, row=self.row_index).value = self.titles['crops_name']
        self.__sheet.cell(column=2, row=self.row_index).value = crops
        self.__crop_set.clear()

    def __set_104y__hire_or_short_hire(self, hire_list, is104y=True):
        if not hire_list:
            return
        if is104y:
            self.titles['fir_half'][0] = '[104農普每月僱工]'
        else:
            self.titles['fir_half'][0] = '[106每月臨時僱工]'

        self.row_index = 2
        self.__set_title('fir_half')
        for index, mon, in enumerate(hire_list[:6], start=2):
            self.__sheet.cell(column=index, row=self.row_index).value = str(mon)

        self.row_index = 1
        self.__set_title('sec_half', 2)
        for index, mon, in enumerate(hire_list[6:], start=2):
            self.__sheet.cell(column=index, row=self.row_index).value = str(mon)

    def __set_hire_lack_or_short_lack(self, data_list, title, is_short_lack=False):
        if not data_list:
            return
        self.row_index = 2
        number_d = {
            'hire': '常僱人數',
            'lack': '常缺人數',
            'short_lack': '臨缺人數',
        }
        work_type = [d['工作類型'] for d in data_list]
        work_type.insert(0, '工作類型')
        number = [d[number_d[title]] for d in data_list]
        number.insert(0, '人數')
        mon = [reduce(lambda a, b: str(a) + ', ' + str(b), l) for l in [d['months'] for d in data_list]]
        mon.insert(0, '月份')
        _list = [work_type, number, mon]
        if is_short_lack:
            product = [d['產品名稱'] for d in data_list]
            product.insert(0, '產品')
            _list.insert(0, product)

        self.__sheet.cell(column=1, row=self.row_index).value = self.titles[title]
        for x, inner_list in enumerate(_list, start=1):
            if x > 1:
                self.row_index = 1
            for index, i in enumerate(inner_list, start=2):
                self.__sheet.cell(column=index, row=self.row_index).value = str(i)
                if len(str(i)) > 8:
                    self.__sheet.cell(column=index, row=self.row_index).alignment = Alignment(wrap_text=True)

    def __set_lack_situation(self, _str):
        if not _str:
            return
        self.row_index = 2
        self.__sheet.cell(column=1, row=self.row_index).value = self.titles['lack_situation']
        self.__sheet.cell(column=2, row=self.row_index).value = _str

    def __set_seprate_symbol(self):
        self.row_index = 1
        self.__sheet.cell(column=1, row=self.row_index).value = ' ' + '=' * 129 + ' '
        self.row_index = 1
        self.__sheet.cell(column=1, row=self.row_index).value = ''

    def set_data(self, data):
        self.__set_sample_base_data(data)
        self.__set_household_data(data['household'])
        self.__set_crop_sbdy_data(data['crop_sbdy'])
        self.__set_disaster_data(data['disaster'])
        self.__set_crops_name()
        self.__set_104y__hire_or_short_hire(data['mon_hire_104y'])
        self.__set_hire_lack_or_short_lack(data['hire_106y'], 'hire')
        self.__set_104y__hire_or_short_hire(data['short_hire_106y'], False)
        self.__set_lack_situation(data['lack_situation'])
        self.__set_hire_lack_or_short_lack(data['lack_106y'], 'lack')
        self.__set_hire_lack_or_short_lack(data['short_lack_106y'], 'short_lack')
        self.__set_seprate_symbol()
        self.__wb.save(os.path.join(self.__path, self.__county + '.xlsx'))