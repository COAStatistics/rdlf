import datetime
import json
import openpyxl
import os
import time
from collections import namedtuple
from generatedata import FILES
from log import log
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


MAIN = False
# SAMPLE_PATH = '..\\..\\input\\easy.txt'
SAMPLE_PATH = '..\\..\\input\\main_107farmerSurvey.txt' if MAIN else '..\\..\\input\\sub_107farmerSurvey.txt'
JSON_PATH = '..\\..\\output\\json\\公務資料.json' if MAIN else '..\\..\\output\\json\\公務資料_備選.json'
# JSON_PATH = '..\\..\\output\\json\\json.json'
FOLDER_NAME = '主選_公務資料' if MAIN else '備選_公務資料'
FOLDER_PATH = '..\\..\\output\\' + datetime.datetime.now().strftime('%Y%m%d_%H%M%S') + FOLDER_NAME

SAMPLE_TITLES = ['農戶編號', '調查姓名', '電話', '地址', '出生年', '原層別', '連結編號']
HOUSEHOLD_TITLES = ['[戶籍檔]', '出生年', '關係', '死亡或除戶', '農保', '老農津貼', '國保給付', '勞保給付', '勞退給付', '農保給付']
TRANSFER_CROP_TITLES = ['[轉作補貼]', '項目', '作物名稱', '金額', '期別']
DISASTER_TITLES = ['[災害]', '項目', '災害', '核定作物', '核定面積', '金額']
SB_SBDY_TITLES = ['[107小大]', '姓名', '大專業農轉契作', '小地主出租給付', '離農獎勵']
LIVESTOCK_TITLES = ['[畜牧資訊]', '年', '調查時間', '畜牧品項', '在養數量', '屠宰數量', '副產品名稱', '副產品數量']
SAMPLE_ROSTER_TITLES = ['序號', '樣本套號 ', '農戶編號', '連結編號 ', '戶長姓名', '電話 ', '地址 ', '層別 ', '經營種類 ', '可耕作地面積', '成功打勾']
SAMPLE_ATTR = [
    'layer',
    'name',
    'tel',
    'addr',
    'county',
    'town',
    'link_num',
    'id',
    'num',
    'main_type',
    'area',
    'sample_num',
]
Sample = namedtuple('Sample', SAMPLE_ATTR)

TYPE_FLAG = '主選' if MAIN else '備選'
ALIGNMENT = Alignment(horizontal='center', vertical='bottom')
SIDE = Side(style='medium')
BORDER = Border(
    top=SIDE,
    bottom=SIDE,
    left=SIDE,
    right=SIDE
)

# sorted by county
sample_dict = {}
official_data = json.loads(open(JSON_PATH, encoding='utf8').read())

if not os.path.isdir(FOLDER_PATH):
    os.mkdir(FOLDER_PATH)


def set_excel_title(sheet, row_index, flag, titles) -> None:
    if flag == 'sample_roster':
        for index, title in enumerate(titles[0], start=1):
            cell = sheet.cell(row_index, index)
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            cell.value = title
            cell.border = BORDER
    else:
        for index, title in enumerate(titles, start=1):
            sheet.cell(column=index, row=row_index).value = title


def read_sample() -> None:
    """
    讀取 sample 檔並使用 dict, key = county : value = 住在這縣市的人
    """
    with open(SAMPLE_PATH, encoding='utf8') as f:
        for line in f:
            sample = Sample._make(line.split('\t'))
            county = sample.county
            _id = sample.id
            if county not in sample_dict:
                county_l = []
                county_l.append(sample)
                sample_dict[county] = county_l
            else:
                sample_dict.get(county).append(sample)


def output_excel() -> None:
    total = 0
    count = 0
    count_sample = len(sample_dict)
    for county, samples in sample_dict.items():
        count += 1
        total += len(samples)
        print(total)
        wb = openpyxl.Workbook()
        col_index = 1
        row_index = 1
        county = county
        town = samples[0].town
        sheet = wb.active
        print(county, '( {} / {})'.format(count, count_sample))

        for sample in samples:
            scholarship = ''
            sb = ''
            farmer_num = sample.num
            crops = []
            sample_data = official_data.get(farmer_num)

            if row_index - 1 == 0:
                width = list(map(lambda x: x * 1.054, [14.29, 9.29, 16.29, 29.29, 9.29, 11.29, 11.29, 11.29, 11.29]))
                for i in range(1, len(width) + 1):
                    sheet.column_dimensions[get_column_letter(i)].width = width[i - 1]
            set_excel_title(sheet, row_index, 'sample', SAMPLE_TITLES)
            row_index += 1
            info = [
                farmer_num, sample_data.get('name'), sample_data.get('telephone'), sample_data.get('address'),
                sample_data.get('birthday'), sample_data.get('layer'), sample_data.get('serial')
            ]
            for index, value in enumerate(info, start=1):
                sheet.cell(column=index, row=row_index).value = value
                sheet.cell(column=index, row=row_index).alignment = Alignment(wrap_text=True)

            row_index += 1
            sheet.cell(column=col_index,
                       row=row_index).value = ' -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------- '

            row_index += 1
            set_excel_title(sheet, row_index, 'household', HOUSEHOLD_TITLES)
            household = sample_data.get('household')
            household.sort(key=lambda x: x[1])

            for person in household:
                row_index += 1
                for index, p_data in enumerate(person, start=2):
                    if index in [5 + 2, 6 + 2, 7 + 2, 8 + 2] and p_data:
                        sheet.cell(column=index, row=row_index).number_format = '#,###,###'
                        p_data = eval(p_data.replace(',', ''))
                    if index == 9 + 2:
                        if person[9]:
                            scholarship += person[9] + ','
                        continue
                    if index == 10 + 2 and person[10] not in sb:
                        sb += person[10]
                        break
                    sheet.cell(column=index, row=row_index).value = p_data
                    sheet.cell(column=index, row=row_index).alignment = Alignment(horizontal='left')

            # 輸出轉作補貼資料，檢查是否有資料
            crop_sbdy = sample_data.get('cropSbdy')
            if crop_sbdy:
                row_index += 2
                crop_d = {}
                for i in crop_sbdy:
                    crop_name = i[0]
                    amount = int(eval(i[1]))
                    if crop_name not in crop_d:
                        crop_d[crop_name] = amount
                    else:
                        crop_d[crop_name] = crop_d.get(crop_name) + amount
                log.info(county, ', ', town, ', ', farmer_num, ', crop_sbdy = ', crop_d)

                item_index = 0
                set_excel_title(sheet, row_index, 'transfer_crop', TRANSFER_CROP_TITLES)

                for k, v in crop_d.items():
                    row_index += 1
                    item_index += 1
                    sheet.cell(column=2, row=row_index).value = item_index
                    sheet.cell(column=2, row=row_index).alignment = Alignment(horizontal='left')
                    sheet.cell(column=3, row=row_index).value = k

                    if len(k) > 8:
                        sheet.cell(column=3, row=row_index).alignment = Alignment(wrap_text=True)

                    if v:
                        sheet.cell(column=4, row=row_index).number_format = '#,###,###'
                    sheet.cell(column=4, row=row_index).value = v
                    sheet.cell(column=4, row=row_index).alignment = Alignment(horizontal='left')
                    sheet.cell(column=5, row=row_index).value = '1'

                    if k not in crops:
                        crops.append(k)

            # 輸出災害補助資料，檢查是否有資料
            disaster = sample_data.get('disaster')
            if disaster:
                row_index += 1
                item_index = 0
                disaster_d = {}
                for i in disaster:
                    data = {}
                    disaster_name = i[0] + '-' + i[1]
                    area = float(i[2])
                    amount = int(i[3])
                    if disaster_name not in disaster_d:
                        data['area'] = area
                        data['amount'] = amount
                    else:
                        data = disaster_d.get(disaster_name)
                        data['area'] = data.get('area') + area
                        data['amount'] = data.get('amount') + amount
                    disaster_d[disaster_name] = data
                log.info(county, ', ', town, ', ', farmer_num, ', disaster = ', disaster_d)

                row_index += 1
                set_excel_title(sheet, row_index, 'disaster', DISASTER_TITLES)

                for k, v in disaster_d.items():
                    row_index += 1
                    item_index += 1
                    sheet.cell(column=2, row=row_index).value = item_index
                    sheet.cell(column=2, row=row_index).alignment = Alignment(horizontal='left')
                    l = k.split('-')
                    sheet.cell(column=3, row=row_index).value = l[0]

                    if len(l[0]) > 8:
                        sheet.cell(column=3, row=row_index).alignment = Alignment(wrap_text=True)
                    sheet.cell(column=4, row=row_index).value = l[1]
                    sheet.cell(column=5, row=row_index).value = v.get('area')
                    sheet.cell(column=5, row=row_index).alignment = Alignment(horizontal='left')

                    if v.get('amount'):
                        sheet.cell(column=6, row=row_index).number_format = '#,###,###'
                    sheet.cell(column=6, row=row_index).value = v.get('amount')
                    sheet.cell(column=6, row=row_index).alignment = Alignment(horizontal='left')

                    if l[1] not in crops:
                        crops.append(l[1])

            # 年度作物
            if crops:
                row_index += 2
                sheet.cell(column=1, row=row_index).value = '[106y-107y作物]'
                sheet.cell(column=2, row=row_index).value = ','.join(crops)
                log.info(county, ', ', town, ', ', farmer_num, ', crops = ', crops)

            row_index += 1
            sheet.cell(column=col_index,
                       row=row_index).value = ' ================================================================================================================================= '
            row_index += 1
            sheet.cell(column=col_index, row=row_index).value = ''

        excel_name = FOLDER_PATH + '\\' + county + '.xlsx' if MAIN else FOLDER_PATH + '\\' + county + '(備選公務檔)' + '.xlsx'
        wb.save(excel_name)


# start_time = time.time()
# read_sample()
# output_excel()
# m, s = divmod(time.time() - start_time, 60)
# print(int(m), 'min', round(s, 1), 'sec')
# log.info(int(m), ' min ', round(s, 1), ' sec')


def read_result_data() -> dict:
    data_dict = {}
    
    for data in json.loads(open(FILES['result_json'], encoding='utf8').read()).values():
        county = data.get('addr')[:3]
        if county not in data_dict:
            data_dict[county] = [data]
        else:
            data_dict.get(county).append(data)
            
    return data_dict


def write_data_to_excel() -> None:
    data_dict = read_result_data()
    for i in data_dict.keys():
        print(i)


if __name__ == '__main__':
    write_data_to_excel()